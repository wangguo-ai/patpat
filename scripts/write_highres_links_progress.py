import json
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font


ROOT = Path(__file__).resolve().parents[1]
SRC_XLSX = ROOT / "secondary_process_duration.xlsx"
URL_RESULTS = ROOT / "tmp_images" / "highres_url_results.json"
OUT_XLSX = ROOT / "secondary_process_duration_with_highres_links_progress.xlsx"


def main() -> None:
    results = json.loads(URL_RESULTS.read_text(encoding="utf-8"))
    ok = [r for r in results if r.get("ok") and r.get("url")]

    shutil.copyfile(SRC_XLSX, OUT_XLSX)
    wb = openpyxl.load_workbook(OUT_XLSX)
    ws = wb.worksheets[0]
    ws.column_dimensions["AO"].width = 16
    ws["AO1"] = "图片"
    ws["AO1"].alignment = Alignment(horizontal="center", vertical="center")

    for r in ok:
        row = int(r["row"])
        cell = ws[f"AO{row}"]
        cell.value = "查看原图"
        cell.hyperlink = r["url"]
        cell.font = Font(color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, 24)

    wb.save(OUT_XLSX)
    screenshot_rows = [r["row"] for r in ok if 531 <= int(r["row"]) <= 566]
    print(
        json.dumps(
            {
                "output": str(OUT_XLSX),
                "links_written": len(ok),
                "rows_min": min((int(r["row"]) for r in ok), default=None),
                "rows_max": max((int(r["row"]) for r in ok), default=None),
                "screenshot_rows_written_count": len(screenshot_rows),
                "screenshot_rows_written": screenshot_rows,
                "file_size": OUT_XLSX.stat().st_size,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
