from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


WB_PATH = Path(r"D:\报销工作台\04_WorkBuddy任务\WorkBuddy采购入库任务表.xlsx")
PURCHASE_PATH = Path(r"D:\Gemini Chrome下载\版料采购单2026年07月16日.xlsx")
BACKUP_DIR = Path(r"D:\报销工作台\04_WorkBuddy任务\backup")
NOW = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
STAMP = datetime.now().strftime("%Y%m%d_%H%M%S")


DOCS = {
    "bintao": r"D:\报销工作台\01_供应商单据\微信图片_20260716095512.png",
    "wensheng": r"D:\报销工作台\01_供应商单据\微信图片_20260716095523.png",
    "yifeng": r"D:\报销工作台\01_供应商单据\微信图片_20260716095043.png",
    "baonianfeng": r"D:\报销工作台\01_供应商单据\微信图片_20260716095024.jpg",
    "longfa": r"D:\报销工作台\01_供应商单据\微信图片_20260716095014.jpg",
    "jiexing": r"D:\报销工作台\01_供应商单据\微信图片_20260715100759.jpg",
}


def item(order_no: str, supplier: str, doc_key: str, doc_no: str, sale_date: str,
         qty: float, unit_price: float, note: str) -> dict:
    amount = round(qty * unit_price, 4)
    return {
        "order_no": order_no,
        "supplier": supplier,
        "doc_path": DOCS[doc_key],
        "doc_no": doc_no,
        "sale_date": sale_date,
        "qty": qty,
        "unit_price": round(unit_price, 4),
        "amount": amount,
        "note": note,
    }


ROWS = [
    # 缤涛 2026/7/15，单据总额 377，全部按 13 元/米写入。
    item("MCG202607140048", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 3, 13, "缤涛阻燃仿棉拉架200G 2#大红，3米，单据13元/米。"),
    item("MCG202607140045", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 3, 13, "缤涛阻燃仿棉拉架200G 49#深卡其，3米，单据13元/米。"),
    item("MCG202607140039", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 3, 13, "缤涛阻燃仿棉拉架200G 3#墨蓝，单据合计8米，按采购单拆3米。"),
    item("MCG202607140025", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 3, 13, "缤涛阻燃仿棉拉架200G 3#墨蓝，单据合计8米，按采购单拆3米。"),
    item("MCG202607140004", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 2, 13, "缤涛阻燃仿棉拉架200G 3#墨蓝，单据合计8米，按采购单拆2米。"),
    item("MCG202607140035", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 3, 13, "缤涛阻燃仿棉拉架200G 1#黑色，单据合计6米，按采购单拆3米。"),
    item("MCG202607140026", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 3, 13, "缤涛阻燃仿棉拉架200G 1#黑色，单据合计6米，按采购单拆3米。"),
    item("MCG202607140049", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 2, 13, "缤涛阻燃1*1拉架罗纹260G 3#墨蓝，单据合计3米，按采购单拆2米。"),
    item("MCG202607140024", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 1, 13, "缤涛阻燃1*1拉架罗纹260G 3#墨蓝，单据合计3米，按采购单拆1米。"),
    item("MCG202607140046", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 1, 13, "缤涛阻燃1*1拉架罗纹260G 49#深卡其，单据合计3米，按采购单拆1米。"),
    item("MCG202607140044", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 2, 13, "缤涛阻燃1*1拉架罗纹260G 49#深卡其，单据合计3米，按采购单拆2米。"),
    item("MCG202607140038", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 1, 13, "缤涛阻燃1*1拉架罗纹260G 1#黑色，单据合计3米，按采购单拆1米。"),
    item("MCG202607140036", "佛山市缤涛纺织有限公司", "bintao", "", "2026-07-15", 2, 13, "缤涛阻燃1*1拉架罗纹260G 1#黑色，单据合计3米，按采购单拆2米。"),

    item("MCG202607140018", "文盛行（原乾辰世景）", "wensheng", "XS136535", "2026-07-14", 2, 12, "文盛行加厚QC320 米白，2米，单据12元/米。"),
    item("MCG202607140008", "文盛行（原乾辰世景）", "wensheng", "XS136535", "2026-07-14", 2, 13, "文盛行6113-a 4#浅绿，2米，单据13元/米。"),

    item("MCG202607100014", "怡丰布行", "yifeng", "XS-260713523", "2026-07-13", 1, 20, "怡丰6-1275 105#浅蓝，单据金额20元，按金额倒核20元/米。"),

    item("MCG202607140016", "宝年丰布业", "baonianfeng", "BNF2607150035", "2026-07-15", 2, 19.95, "宝年丰6032#32S棉拉架22#宝蓝，供应商按1KG/39.90元开单，采购按2米入库，折19.95元/米。"),
    item("MCG202607140001", "宝年丰布业", "baonianfeng", "BNF2607150035", "2026-07-15", 1, 18.45, "宝年丰6032#32S棉拉架40#橙红，供应商按0.5KG*36.90=18.45元开单，采购按1米入库。"),

    item("MCG202607150006", "隆发罗纹", "longfa", "XS202607150006", "2026-07-15", 1, 30, "隆发2*2OE棉21支 1363宝蓝，1米，单据30元。"),
    item("MCG202607150002", "隆发罗纹", "longfa", "XS202607150006", "2026-07-15", 1, 25, "隆发2*2棉21支 629彩蓝，1米，单据25元。"),

    item("MCG202607130012", "杰兴（杰诺）布行", "jiexing", "BD-20260714-0236", "2026-07-14", 1, 25, "杰兴32支棉拉架 74#浅宝蓝，1米，单据25元。"),
]


def headers(ws):
    return {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}


def first_blank_row(ws, key_col: int) -> int:
    for r in range(2, ws.max_row + 2):
        if ws.cell(r, key_col).value in (None, ""):
            return r
    return ws.max_row + 1


def set_row(ws, row_num: int, h: dict[str, int], values: dict[str, object]) -> None:
    for key, value in values.items():
        ws.cell(row_num, h[key]).value = value


def load_purchase() -> dict[str, dict]:
    wb = load_workbook(PURCHASE_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    h = headers(ws)
    out = {}
    for r in range(2, ws.max_row + 1):
        order_no = ws.cell(r, h["版料采购单号"]).value
        if not order_no:
            continue
        out[str(order_no)] = {
            "supplier": ws.cell(r, h["供应商名称"]).value,
            "material": ws.cell(r, h["供应商物料编号"]).value,
            "color": ws.cell(r, h["供应商物料颜色"]).value,
            "color_no": ws.cell(r, h["供应商物料色号"]).value,
            "unit": ws.cell(r, h["采购单位"]).value,
            "demand_qty": ws.cell(r, h["需求数量"]).value,
        }
    return out


def main() -> None:
    counts = Counter(row["order_no"] for row in ROWS)
    duplicates = [order for order, count in counts.items() if count > 1]
    if duplicates:
        raise SystemExit(f"duplicate rows in plan: {duplicates}")

    purchase = load_purchase()
    missing = [row["order_no"] for row in ROWS if row["order_no"] not in purchase]
    if missing:
        raise SystemExit(f"purchase rows missing: {missing}")

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / f"WorkBuddy采购入库任务表_backup_20260716_{STAMP}.xlsx"
    copy2(WB_PATH, backup)

    wb = load_workbook(WB_PATH)
    task_ws = wb["入库任务"]
    detail_ws = wb["影刀操作明细"]
    archive_ws = wb["已经执行完的任务"]
    task_h = headers(task_ws)
    detail_h = headers(detail_ws)
    archive_h = headers(archive_ws)

    existing = set()
    for ws, h in [(detail_ws, detail_h), (archive_ws, archive_h)]:
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, h["采购单号"]).value
            if val:
                existing.add(str(val))
    already = sorted({row["order_no"] for row in ROWS if row["order_no"] in existing})
    if already:
        raise SystemExit(f"orders already in WorkBuddy detail/archive: {already}")

    task_row = first_blank_row(task_ws, task_h["采购单号列表"])
    detail_row = first_blank_row(detail_ws, detail_h["采购单号"])

    for row in ROWS:
        purchase_row = purchase[row["order_no"]]
        task_id = f"WB-IN-20260716-XW-{row['order_no']}"
        set_row(task_ws, task_row, task_h, {
            "任务ID": task_id,
            "任务类型": "采购入库+打印标签",
            "供应商": row["supplier"],
            "供应商单据路径": row["doc_path"],
            "供应商单据号": row["doc_no"],
            "销售日期": row["sale_date"],
            "采购单号列表": row["order_no"],
            "总入库数量": row["qty"],
            "总金额": row["amount"],
            "是否已人工确认": "是",
            "是否允许入库保存": "是",
            "是否需要打印标签": "是",
            "任务状态": "待执行",
            "执行要求": "影刀执行前核对页面明细；按本行数量、单价执行采购完成。",
            "创建时间": NOW,
            "备注": "2026-07-16 小王逐张核对供应商单据与采购单后写入当前影刀操作明细。",
            "采购单号匹配规则": "供应商单据原图/2026-07-16采购单导出匹配",
            "图灵采购查询入口": "https://turing.patpat.shop/mrp/supply-chain/surface-accessories-m-r-p/plate-material-management/purchase-manage-list",
            "是否需图灵实时匹配": "否",
        })
        set_row(detail_ws, detail_row, detail_h, {
            "任务ID": task_id,
            "采购单号": row["order_no"],
            "采购状态": "待入库",
            "供应商": row["supplier"],
            "物料": purchase_row["material"],
            "颜色": purchase_row["color"],
            "色号": purchase_row["color_no"],
            "单位": purchase_row["unit"],
            "需求数量": purchase_row["demand_qty"],
            "本次入库数量": row["qty"],
            "实际单价": row["unit_price"],
            "本次金额": row["amount"],
            "标签打印": "待确认",
            "核对备注": row["note"],
            "采购单号来源": "供应商单据原图/2026-07-16采购单导出匹配",
            "采购单号匹配状态": "已匹配",
            "缺失时匹配动作": "影刀执行前核对页面明细；按本行数量、单价执行采购完成。",
            "任务状态": "待执行",
        })
        task_row += 1
        detail_row += 1

    wb.save(WB_PATH)

    totals = defaultdict(float)
    for row in ROWS:
        totals[row["supplier"]] += row["amount"]
    print(f"backup={backup}")
    print(f"rows={len(ROWS)}")
    print(f"total={sum(row['amount'] for row in ROWS):.4f}")
    for supplier, total in sorted(totals.items()):
        print(f"supplier_total\t{supplier}\t{total:.4f}")


if __name__ == "__main__":
    main()
