import openpyxl, shutil, json
from datetime import datetime

def U(s): return s.encode('ascii').decode('unicode_escape')
base = "D:" + "\\" + U("\\u62a5\\u9500\\u5de5\\u4f5c\\u53f0") + "\\04_WorkBuddy" + U("\\u4efb\\u52a1")
path = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868.xlsx")
backup = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868_backup_before_fill_pending_mcg_20260623_") + datetime.now().strftime('%H%M%S') + ".xlsx"
shutil.copy2(path, backup)
wb=openpyxl.load_workbook(path)
ws=wb.worksheets[1]
updates={
    'WB-IN-20260623-XW-PENDING-BINTAO-260427004': {
        'task':'WB-IN-20260623-XW-MCG202606220001',
        'order':'MCG202606220001',
        'note':'来源供应商单据: 缤涛成品销售出货单 260427004；用户确认对应采购单号 MCG202606220001；BT0015A 2#大红，2米，13元/米，金额26。'
    },
    'WB-IN-20260623-XW-PENDING-HUAXUN-PL2606220536': {
        'task':'WB-IN-20260623-XW-MCG202606170002',
        'order':'MCG202606170002',
        'note':'来源供应商单据: 华讯昌达 STE26062201969\\PL2606220536；用户确认对应采购单号 MCG202606170002；仿棉2*2罗纹-RS15#丁香紫，0.9KG，37元/KG，单据金额33元（供应商取整）。'
    }
}
updated=[]
for r in range(2, ws.max_row+1):
    task=ws.cell(r,1).value
    if task in updates:
        u=updates[task]
        ws.cell(r,1).value=u['task']
        ws.cell(r,2).value=u['order']
        ws.cell(r,3).value='待入库'
        ws.cell(r,14).value=u['note']
        ws.cell(r,16).value='已匹配'
        ws.cell(r,17).value='影刀执行前核对页面明细；按本行数量、单价执行采购完成。'
        ws.cell(r,18).value='待执行'
        updated.append(u['order'])
wb.save(path)
print(json.dumps({'xlsx':path,'backup':backup,'updated':updated}, ensure_ascii=False, indent=2))
