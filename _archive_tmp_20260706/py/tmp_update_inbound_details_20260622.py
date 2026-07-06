import openpyxl, json, shutil
from datetime import datetime


def U(s):
    return s.encode('ascii').decode('unicode_escape')

base = "D:" + "\\" + U("\\u62a5\\u9500\\u5de5\\u4f5c\\u53f0") + "\\04_WorkBuddy" + U("\\u4efb\\u52a1")
xlsx_path = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868.xlsx")
ready_path = "D:" + "\\ai" + U("\\u5171\\u4eab\\u76d8") + "\\MyBrain\\xiaowang_completion_ready_20260622.json"
backup_path = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868_backup_before_all_20260622_") + datetime.now().strftime('%H%M%S') + ".xlsx"

orders = '''MCG202606180024
MCG202606180015
MCG202606180003
MCG202606180001
MCG202606170001
MCG202606160044
MCG202606180010
MCG202606180009
MCG202606180023
MCG202606180013
MCG202606180012
MCG202606170017
MCG202606170020
MCG202606170018
MCG202606170022
MCG202606170012
MCG202606170013
MCG202606170014
MCG202606170015
MCG202606180021
MCG202606180018
MCG202606170031
MCG202606180017
MCG202606170033
MCG202606170034
MCG202606170035'''.splitlines()

with open(ready_path, 'r', encoding='utf-8') as f:
    ready = json.load(f)
ready_map = {r['order_no']: r for r in ready}
completed = {'MCG202606180021', 'MCG202606180018'}
hold = {
    'MCG202606180024': ('怡丰布行', '怡丰布行 XS-260618637', '缺少可靠到单号的实际数量/单价映射，需人工确认后再批量'),
    'MCG202606180015': ('怡丰布行', '怡丰布行 XS-260618637', '缺少可靠到单号的实际数量/单价映射，需人工确认后再批量'),
    'MCG202606180003': ('怡丰布行', '怡丰布行 XS-260618637', '缺少可靠到单号的实际数量/单价映射，需人工确认后再批量'),
    'MCG202606180001': ('怡丰布行', '怡丰布行 XS-260618637', '缺少可靠到单号的实际数量/单价映射，需人工确认后再批量'),
    'MCG202606170001': ('怡丰布行', '怡丰布行 XS-260617529', '缺少可靠到单号的实际数量/单价映射，需人工确认后再批量'),
    'MCG202606160044': ('怡丰布行', '怡丰布行 XS-260616481', '缺少可靠到单号的实际数量/单价映射，需人工确认后再批量'),
    'MCG202606180010': ('汕头明达纺织', '汕头明达纺织 XS20276966', '供应商单据单位/图灵采购单位口径需确认，暂不自动提交'),
    'MCG202606180009': ('汕头明达纺织', '汕头明达纺织 XS20276966', '供应商单据单位/图灵采购单位口径需确认，暂不自动提交'),
    'MCG202606170031': ('优粤纺织', '优粤纺织 UK92039-56#暗紫', '前面核对出现过 U5919-73#湖兰 与 UK92039-56#暗紫 两种对应；当前清单按 UK92039-56#暗紫 2米 单价25，需确认后执行'),
}

def supplier_for(file_name):
    if '优粤纺织' in file_name:
        return '优粤纺织'
    if '文盛行' in file_name:
        return '文盛行（原乾辰世景）'
    if '佛山市缤涛纺织' in file_name:
        return '佛山市缤涛纺织有限公司'
    return ''

def row_for(order):
    task_id = 'WB-IN-20260622-XW-' + order
    if order in completed:
        return [task_id, order, '待入库', '宝年丰布业', '', '', '', '米', '', 1, 17.05, 17.05, '待确认', '来源供应商单据: 宝年丰布业 BNF2606180279；采购完成提交已测试跑通，标签打印仅尝试点击未验证出纸', '供应商单据OCR', '已匹配', '小吴入库前核对页面明细；标签打印需人工确认', '待执行']
    if order in hold:
        supplier, source, reason = hold[order]
        rec = ready_map.get(order, {})
        qty = rec.get('quantity', '')
        price = rec.get('unit_price', '')
        amount = qty * price if isinstance(qty, (int, float)) and isinstance(price, (int, float)) else ''
        unit = '米' if qty else ''
        return [task_id, order, '待确认', supplier, '', '', '', unit, '', qty, price, amount, '待确认', '来源供应商单据: ' + source + '；' + reason, '供应商单据/人工核对', '待确认', '暂不入库保存，先补充数量/单价/对应关系确认', '待确认']
    rec = ready_map[order]
    qty = rec.get('quantity', '')
    price = rec.get('unit_price', '')
    amount = qty * price if isinstance(qty, (int, float)) and isinstance(price, (int, float)) else ''
    file_name = rec.get('file', '')
    return [task_id, order, '待入库', supplier_for(file_name), '', '', '', '米', '', qty, price, amount, '待确认', '来源供应商单据: ' + file_name + '；' + rec.get('raw_text',''), '供应商单据OCR', '已匹配', '小王已准备采购完成数据，交小吴入库前仍需核对页面明细', '待执行']

shutil.copy2(xlsx_path, backup_path)
wb = openpyxl.load_workbook(xlsx_path)
ws = wb.worksheets[1]
existing = {}
for r in range(2, ws.max_row + 1):
    order = ws.cell(r, 2).value
    if order:
        existing[order] = r
added, updated = [], []
for order in orders:
    vals = row_for(order)
    if order in existing:
        r = existing[order]
        for c, v in enumerate(vals, start=1):
            ws.cell(r, c).value = v
        updated.append(order)
    else:
        ws.append(vals)
        added.append(order)
wb.save(xlsx_path)
print(json.dumps({'xlsx': xlsx_path, 'backup': backup_path, 'added': added, 'updated': updated, 'total': len(orders)}, ensure_ascii=False, indent=2))
