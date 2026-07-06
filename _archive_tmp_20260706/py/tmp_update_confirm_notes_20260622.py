import openpyxl, shutil, json
from datetime import datetime

def U(s): return s.encode('ascii').decode('unicode_escape')
base = "D:" + "\\" + U("\\u62a5\\u9500\\u5de5\\u4f5c\\u53f0") + "\\04_WorkBuddy" + U("\\u4efb\\u52a1")
path = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868.xlsx")
backup = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868_backup_before_confirm_notes_20260622_") + datetime.now().strftime('%H%M%S') + ".xlsx"
shutil.copy2(path, backup)
wb = openpyxl.load_workbook(path)
ws = wb.worksheets[1]

yifeng = {'MCG202606180024','MCG202606180015','MCG202606180003','MCG202606180001','MCG202606170001','MCG202606160044'}
mingda = {'MCG202606180010','MCG202606180009'}
updated=[]
for r in range(2, ws.max_row + 1):
    order = ws.cell(r, 2).value
    if order in yifeng:
        ws.cell(r, 14).value = (ws.cell(r, 14).value or '').split('；缺少可靠')[0] + '；用户已确认差异原因：怡丰部分物料单价优惠10元；仍需补齐本采购单实际数量/实际单价后再执行。'
        ws.cell(r, 16).value = '差异原因已确认，待补数量单价'
        ws.cell(r, 17).value = '暂不入库保存，先补齐每个采购单的实际数量/实际单价。'
        ws.cell(r, 18).value = '待确认'
        updated.append(order)
    elif order in mingda:
        ws.cell(r, 14).value = (ws.cell(r, 14).value or '').split('；供应商单据单位')[0] + '；用户已确认差异原因：供应商做了四舍五入；仍需补齐本采购单实际数量/实际单价后再执行。'
        ws.cell(r, 16).value = '差异原因已确认，待补数量单价'
        ws.cell(r, 17).value = '暂不入库保存，先补齐每个采购单的实际数量/实际单价。'
        ws.cell(r, 18).value = '待确认'
        updated.append(order)
    elif order == 'MCG202606170031':
        ws.cell(r, 13).value = '待确认'
        ws.cell(r, 14).value = '来源供应商单据: 优粤纺织 UK92039-56#暗紫；用户已确认对应无误：2米，单价25，金额50。'
        ws.cell(r, 16).value = '已匹配'
        ws.cell(r, 17).value = '小王已准备采购完成数据，交小吴入库前仍需核对页面明细。'
        ws.cell(r, 18).value = '待执行'
        updated.append(order)
wb.save(path)
print(json.dumps({'xlsx': path, 'backup': backup, 'updated': updated}, ensure_ascii=False, indent=2))
