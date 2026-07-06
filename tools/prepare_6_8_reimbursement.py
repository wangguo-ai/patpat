from __future__ import annotations

from collections import defaultdict
import argparse
from decimal import Decimal, ROUND_HALF_UP
import os
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SOURCE_XLSX = Path("D:/Gemini Chrome下载/6.8报销.xlsx")
REFERENCE_XLSX = Path("D:/app/RPA/影刀数据表格.xlsx")
BASE_DIR = Path("D:/ai共享盘/MyBrain")
VOUCHER_DIR = BASE_DIR / "支付凭证"
OUT_DIR = BASE_DIR / "MyBrain"
OUT_XLSX = OUT_DIR / "6.8报销_影刀格式.xlsx"
LOG_TXT = OUT_DIR / "6.8凭证改名记录_按原表序号.txt"

VOUCHER_RE = re.compile(r"^(?P<amount>\d+(?:\.\d+)?)(?:-(?P<idx>\d+))?-(?P<supplier>.+)\.(?P<ext>[^.]+)$")
AMOUNT_TOLERANCE = Decimal("0.50")


def dec(value: Any) -> Decimal:
    return Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def amount_label(value: Decimal) -> str:
    amount = value.quantize(Decimal("0.01"))
    if amount == amount.to_integral_value():
        return str(int(amount))
    return format(amount.normalize(), "f")


def norm(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .replace("（", "(")
        .replace("）", ")")
        .replace("巨鍍", "巨鍑")
        .replace("巨凳", "巨鍑")
    )


def sort_key(value: Any) -> tuple[int, Decimal | str]:
    try:
        return (0, Decimal(str(value)))
    except Exception:
        return (1, str(value))


def supplier_score(voucher_supplier: str, source_supplier: str) -> int:
    voucher = norm(voucher_supplier)
    source = norm(source_supplier)
    if not voucher or not source:
        return 0
    if voucher == source:
        return 100
    if voucher in source or source in voucher:
        return 80
    if len(voucher) >= 2 and voucher[:2] == source[:2]:
        return 50
    return 0


def load_source_groups() -> list[tuple[Any, list[dict[str, Any]]]]:
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    worksheet = workbook.active
    groups: dict[Any, list[dict[str, Any]]] = defaultdict(list)

    for source_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        serial = row[0]
        if serial is None or str(serial).strip() == "":
            continue
        diff = dec(row[27] if row[27] is not None else 0)
        amount = dec(row[28]) if row[28] is not None else (dec(row[26]) + diff).quantize(Decimal("0.01"))
        groups[serial].append(
            {
                "source_row": source_row,
                "serial": serial,
                "po": row[1],
                "supplier": row[9] or "",
                "stock_amount": amount,
                "payment_amount": amount,
                "date": row[42] or row[37] or "",
                "diff": float(diff) if diff != diff.to_integral_value() else int(diff),
                "system_amount": row[26] if row[26] is not None else amount,
                "remark": row[6] or "",
            }
        )

    return [(serial, groups[serial]) for serial in sorted(groups.keys(), key=sort_key)]


def group_total(rows: list[dict[str, Any]]) -> Decimal:
    return sum((row["payment_amount"] for row in rows), Decimal("0.00")).quantize(Decimal("0.01"))


def parse_vouchers() -> list[dict[str, Any]]:
    vouchers = []
    for path in sorted(VOUCHER_DIR.glob("*"), key=lambda item: item.name):
        if not path.is_file():
            continue
        match = VOUCHER_RE.match(path.name)
        if match:
            amount = dec(match.group("amount"))
            supplier = match.group("supplier")
            idx = int(match.group("idx") or 1)
        else:
            amount = None
            supplier = path.stem
            idx = 1
        vouchers.append(
            {
                "path": path,
                "amount": amount,
                "supplier": supplier,
                "idx": idx,
                "ext": path.suffix.lower(),
                "planned_used": False,
                "renamed_used": False,
            }
        )
    return vouchers


def next_upload_name(amount: Decimal, supplier: str, counters: dict[tuple[str, str], int], reserved: set[str]) -> str:
    key = (amount_label(amount), supplier)
    while True:
        counters[key] += 1
        if counters[key] == 1:
            name = f"{amount_label(amount)}-{supplier}"
        else:
            name = f"{amount_label(amount)}-{counters[key]}-{supplier}"
        if name not in reserved:
            reserved.add(name)
            return name


def solve_supplier_run(items: list[dict[str, Any]], vouchers: list[dict[str, Any]]) -> list[dict[str, Any]]:
    supplier = items[0]["supplier"]
    candidates = tuple(
        index
        for index, voucher in enumerate(vouchers)
        if not voucher["planned_used"]
        and voucher["amount"] is not None
        and supplier_score(voucher["supplier"], supplier)
    )
    memo: dict[tuple[int, tuple[int, ...]], tuple[int, int, int, list[dict[str, Any]]]] = {}

    def recurse(position: int, remaining: tuple[int, ...]) -> tuple[int, int, int, list[dict[str, Any]]]:
        key = (position, remaining)
        if key in memo:
            return memo[key]
        if position >= len(items):
            return (0, len(remaining), 0, [])

        best: tuple[int, int, int, list[dict[str, Any]]] | None = None
        total = Decimal("0.00")
        for end in range(position, len(items)):
            total += items[end]["total"]
            matches = []
            for index in remaining:
                amount_delta = abs(vouchers[index]["amount"] - total)
                tolerance = Decimal("1.00") if "明达" in norm(supplier) else AMOUNT_TOLERANCE
                if amount_delta <= tolerance:
                    matches.append((amount_delta, index))
            for amount_delta, voucher_index in sorted(matches):
                next_remaining = tuple(index for index in remaining if index != voucher_index)
                miss, unused, chunks, plan = recurse(end + 1, next_remaining)
                candidate = (
                    miss,
                    unused,
                    chunks + 1,
                    [
                        {
                            "start": position,
                            "end": end,
                            "voucher_index": voucher_index,
                            "amount": vouchers[voucher_index]["amount"],
                            "source_total": total,
                            "amount_delta": amount_delta,
                            "missing": False,
                        }
                    ]
                    + plan,
                )
                if best is None or candidate[:3] < best[:3]:
                    best = candidate

        total = items[position]["total"]
        miss, unused, chunks, plan = recurse(position + 1, remaining)
        candidate = (
            miss + 1,
            unused,
            chunks + 1,
            [
                {
                    "start": position,
                    "end": position,
                    "voucher_index": None,
                    "amount": total,
                    "source_total": total,
                    "amount_delta": Decimal("0.00"),
                    "missing": True,
                }
            ]
            + plan,
        )
        if best is None or candidate[:3] < best[:3]:
            best = candidate

        memo[key] = best
        return best

    return recurse(0, candidates)[3]


def plan_upload_names(groups: list[tuple[Any, list[dict[str, Any]]]], vouchers: list[dict[str, Any]]) -> tuple[dict[Any, str], list[list[Any]]]:
    upload_names: dict[Any, str] = {}
    assignment_details: list[list[Any]] = []
    counters: dict[tuple[str, str], int] = defaultdict(int)
    reserved: set[str] = set()

    summaries = [
        {
            "serial": serial,
            "rows": rows,
            "supplier": rows[0]["supplier"] or rows[-1]["supplier"],
            "total": group_total(rows),
        }
        for serial, rows in groups
    ]

    run: list[dict[str, Any]] = []
    for item in summaries + [None]:
        if item is not None and (not run or norm(item["supplier"]) == norm(run[-1]["supplier"])):
            run.append(item)
            continue
        if run:
            plan = solve_supplier_run(run, vouchers)
            for chunk in plan:
                chunk_items = run[chunk["start"] : chunk["end"] + 1]
                supplier = chunk_items[0]["supplier"]
                upload_name = next_upload_name(chunk["amount"], supplier, counters, reserved)
                last_serial = chunk_items[-1]["serial"]
                upload_names[last_serial] = upload_name
                if chunk["voucher_index"] is not None:
                    vouchers[chunk["voucher_index"]]["planned_used"] = True
                assignment_details.append(
                    [
                        ",".join(str(x["serial"]) for x in chunk_items),
                        upload_name,
                        float(chunk["amount"]),
                        supplier,
                        "缺少支付凭证"
                        if chunk["missing"]
                        else (
                            f"已分配支付凭证；原表合计{amount_label(chunk['source_total'])}，凭证{amount_label(chunk['amount'])}"
                            if chunk["amount_delta"]
                            else "已分配支付凭证"
                        ),
                    ]
                )
        run = [] if item is None else [item]

    return upload_names, assignment_details


def build_rows(groups: list[tuple[Any, list[dict[str, Any]]]], upload_names: dict[Any, str], assignment_details: list[list[Any]]) -> tuple[list[list[Any]], list[list[Any]]]:
    table_rows = []
    detail_rows = []
    for serial, rows in groups:
        upload_name = upload_names.get(serial, "")
        for index, row in enumerate(rows):
            excel_row = len(table_rows) + 2
            table_rows.append(
                [
                    row["serial"],
                    row["po"],
                    row["supplier"],
                    upload_name if index == len(rows) - 1 else "",
                    float(row["stock_amount"]),
                    row["date"],
                    f"=E{excel_row}+H{excel_row}",
                    row["diff"],
                    row["system_amount"],
                    "待上传",
                    row["remark"],
                ]
            )
        detail_rows.append(
            [
                serial,
                upload_name,
                len(rows),
                float(sum((row["stock_amount"] for row in rows), Decimal("0.00"))),
                rows[0]["supplier"],
                ",".join(str(row["po"]) for row in rows),
            ]
        )
    detail_rows.append(["", "", "", "", "", ""])
    detail_rows.extend(assignment_details)
    return table_rows, detail_rows


def sync_assignment_statuses(detail_rows: list[list[Any]], rename_log: list[list[Any]]) -> None:
    status_by_name = {}
    for _serial, _source, target_name, result in rename_log:
        if not target_name:
            continue
        upload_name = Path(str(target_name)).stem
        if result in {"已改名", "无需改名"}:
            status_by_name[upload_name] = "已分配支付凭证"
        elif result == "缺少支付凭证":
            status_by_name[upload_name] = "缺少支付凭证"

    for row in detail_rows:
        if len(row) == 5 and row[1] in status_by_name:
            row[4] = status_by_name[row[1]]


def find_voucher(vouchers: list[dict[str, Any]], upload_name: str) -> dict[str, Any] | None:
    match = VOUCHER_RE.match(f"{upload_name}.jpg")
    if not match:
        return None
    target_amount = dec(match.group("amount"))
    target_supplier = match.group("supplier")
    choices = []
    for voucher in vouchers:
        if voucher["renamed_used"] or voucher["amount"] != target_amount:
            continue
        score = supplier_score(voucher["supplier"], target_supplier)
        if score:
            choices.append((score, voucher["idx"], len(voucher["supplier"]), voucher))
    if not choices:
        return None
    choices.sort(key=lambda item: (item[0], -item[1], item[2]), reverse=True)
    return choices[0][3]


def rename_vouchers(upload_names: dict[Any, str]) -> list[list[Any]]:
    vouchers = parse_vouchers()
    log = []
    operations = []

    for serial in sorted(upload_names.keys(), key=sort_key):
        upload_name = upload_names[serial]
        voucher = find_voucher(vouchers, upload_name)
        target = VOUCHER_DIR / f"{upload_name}.jpg"
        if voucher is None:
            log.append([serial, "", target.name, "缺少支付凭证"])
            continue
        voucher["renamed_used"] = True
        source = voucher["path"]
        if source.name == target.name:
            log.append([serial, source.name, target.name, "无需改名"])
            continue
        temp = source.with_name(f".__tmp_rename_{os.getpid()}_{len(operations)}{source.suffix}")
        source.rename(temp)
        operations.append((serial, temp, target, source.name))

    for serial, temp, target, original_name in operations:
        if target.exists():
            log.append([serial, original_name, target.name, "失败：目标文件已存在"])
            temp.rename(VOUCHER_DIR / original_name)
            continue
        temp.rename(target)
        log.append([serial, original_name, target.name, "已改名"])

    for voucher in vouchers:
        if not voucher["renamed_used"] and voucher["path"].exists():
            log.append(["", voucher["path"].name, "", "未被原表序号使用"])

    return log


def style_sheet(worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    for col in range(1, worksheet.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for cell in worksheet[letter]:
            width = max(width, len(str(cell.value)) if cell.value is not None else 0)
        worksheet.column_dimensions[letter].width = min(width + 2, 48)


def write_workbook(table_rows: list[list[Any]], detail_rows: list[list[Any]], rename_log: list[list[Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["序号", "采购单号", "供应商", "待上传的支付凭证名", "入库金额", "日期", "金额", "差异", "系统采购金额", "状态", "备注"])
    for row in table_rows:
        sheet.append(row)

    detail_sheet = workbook.create_sheet("序号分组明细")
    detail_sheet.append(["序号/序号组合", "待上传的支付凭证名", "采购行数", "分组金额合计", "供应商", "采购单号/状态"])
    for row in detail_rows:
        detail_sheet.append(row)

    rename_sheet = workbook.create_sheet("改名记录")
    rename_sheet.append(["序号", "原文件名", "新文件名", "结果"])
    for row in rename_log:
        rename_sheet.append(row)

    for worksheet in workbook.worksheets:
        style_sheet(worksheet)

    if REFERENCE_XLSX.exists():
        reference = load_workbook(REFERENCE_XLSX)
        reference_sheet = reference.active
        for col in range(1, min(sheet.max_column, reference_sheet.max_column) + 1):
            letter = get_column_letter(col)
            sheet.column_dimensions[letter].width = reference_sheet.column_dimensions[letter].width

    OUT_DIR.mkdir(exist_ok=True)
    workbook.save(OUT_XLSX)


def write_log(rename_log: list[list[Any]]) -> None:
    with LOG_TXT.open("w", encoding="utf-8") as handle:
        for row in rename_log:
            handle.write("\t".join(str(value) for value in row) + "\n")


def configure_from_args() -> None:
    global SOURCE_XLSX, VOUCHER_DIR, OUT_DIR, OUT_XLSX, LOG_TXT
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=str(SOURCE_XLSX))
    parser.add_argument("--voucher-dir", default=str(VOUCHER_DIR))
    parser.add_argument("--out-dir", default=str(OUT_DIR))
    parser.add_argument("--prefix", default="6.8报销")
    args = parser.parse_args()

    SOURCE_XLSX = Path(args.source)
    VOUCHER_DIR = Path(args.voucher_dir)
    OUT_DIR = Path(args.out_dir)
    OUT_XLSX = OUT_DIR / f"{args.prefix}_影刀格式.xlsx"
    LOG_TXT = OUT_DIR / f"{args.prefix}凭证改名记录_按原表序号.txt"


def main() -> None:
    configure_from_args()
    OUT_DIR.mkdir(exist_ok=True)
    groups = load_source_groups()
    planned_vouchers = parse_vouchers()
    upload_names, assignment_details = plan_upload_names(groups, planned_vouchers)
    table_rows, detail_rows = build_rows(groups, upload_names, assignment_details)
    rename_log = rename_vouchers(upload_names)
    sync_assignment_statuses(detail_rows, rename_log)
    write_workbook(table_rows, detail_rows, rename_log)
    write_log(rename_log)

    statuses = defaultdict(int)
    for row in rename_log:
        statuses[row[3]] += 1
    merged = [row for row in assignment_details if "," in str(row[0]) and row[4] == "已分配支付凭证"]
    print(f"OUT_XLSX={OUT_XLSX}")
    print(f"LOG_TXT={LOG_TXT}")
    print(f"GROUPS={len(groups)}")
    print(f"TABLE_ROWS={len(table_rows)}")
    print(f"ASSIGNMENTS={len(assignment_details)}")
    print(f"RENAMED={statuses['已改名']}")
    print(f"UNCHANGED={statuses['无需改名']}")
    print(f"MISSING={statuses['缺少支付凭证']}")
    print(f"UNUSED={statuses['未被原表序号使用']}")
    print(f"MERGED={len(merged)}")
    for row in merged:
        print(f"MERGED_DETAIL={row[0]} -> {row[1]}")


if __name__ == "__main__":
    main()
