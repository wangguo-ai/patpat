from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
import os
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path("D:/ai共享盘/MyBrain")
SOURCE_XLSX = Path("D:/Gemini Chrome下载/6.8报销.xlsx")
REFERENCE_XLSX = Path("D:/app/RPA/影刀数据表格.xlsx")
VOUCHER_DIR = BASE / "支付凭证"
OUT_DIR = BASE / "MyBrain"
OUT_XLSX = OUT_DIR / "6.8报销_影刀格式.xlsx"
LOG_TXT = OUT_DIR / "6.8凭证改名记录_按原表序号.txt"

VOUCHER_RE = re.compile(r"^(?P<amount>\d+(?:\.\d+)?)(?:-(?P<idx>\d+))?-(?P<supplier>.+)\.(?P<ext>[^.]+)$")


def dec(value):
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def amount_label(value):
    value = Decimal(value).quantize(Decimal("0.01"))
    if value == value.to_integral_value():
        return str(int(value))
    return format(value.normalize(), "f")


def norm(value):
    return str(value or "").replace("發", "发").replace("（", "(").replace("）", ")")


def sort_key(value):
    try:
        return (0, Decimal(str(value)))
    except Exception:
        return (1, str(value))


def load_source_groups():
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    worksheet = workbook.active
    groups = defaultdict(list)

    for source_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        serial = row[0]
        if serial is None:
            continue
        amount = dec(row[28] if row[28] is not None else 0)
        groups[serial].append(
            {
                "source_row": source_row,
                "serial": serial,
                "po": row[1],
                "supplier": row[9] or "",
                "stock_amount": amount,
                "date": row[42] or row[37] or "",
                "diff": row[27] if row[27] is not None else 0,
                "system_amount": row[26] if row[26] is not None else amount,
                "remark": row[6] or "",
            }
        )

    return [(serial, groups[serial]) for serial in sorted(groups.keys(), key=sort_key)]


def group_total(rows):
    return sum((row["stock_amount"] for row in rows), Decimal("0.00")).quantize(Decimal("0.01"))


def make_missing_name(amount, supplier, seen_missing, reserved_names):
    key = (amount, supplier)
    while True:
        seen_missing[key] += 1
        if seen_missing[key] == 1:
            name = f"{amount_label(amount)}-{supplier}"
        else:
            name = f"{amount_label(amount)}-{seen_missing[key]}-{supplier}"
        if name not in reserved_names:
            reserved_names.add(name)
            return name


def solve_supplier_run(items, vouchers):
    supplier = items[0]["supplier"]
    voucher_indices = tuple(
        index
        for index, voucher in enumerate(vouchers)
        if not voucher["planned_used"] and voucher["amount"] is not None and supplier_score(voucher["supplier"], supplier)
    )
    memo = {}

    def recurse(position, remaining):
        key = (position, remaining)
        if key in memo:
            return memo[key]
        if position >= len(items):
            return (0, len(remaining), 0, [])

        best = None
        total = Decimal("0.00")
        for end in range(position, len(items)):
            total += items[end]["total"]
            matching = [index for index in remaining if vouchers[index]["amount"] == total]
            for voucher_index in matching:
                next_remaining = tuple(index for index in remaining if index != voucher_index)
                miss, unused, chunks, plan = recurse(end + 1, next_remaining)
                candidate = (
                    miss,
                    unused,
                    chunks + 1,
                    [
                        {
                            "start": position,
                            "end": end,
                            "voucher_index": voucher_index,
                            "upload_name": vouchers[voucher_index]["path"].stem,
                            "missing": False,
                            "amount": total,
                        }
                    ]
                    + plan,
                )
                if best is None or candidate[:3] < best[:3]:
                    best = candidate

        # Missing-voucher fallback is intentionally one source serial only so
        # uncertain merges are driven by actual files in the payment folder.
        total = items[position]["total"]
        miss, unused, chunks, plan = recurse(position + 1, remaining)
        candidate = (
            miss + 1,
            unused,
            chunks + 1,
            [
                {
                    "start": position,
                    "end": position,
                    "voucher_index": None,
                    "upload_name": None,
                    "missing": True,
                    "amount": total,
                }
            ]
            + plan,
        )
        if best is None or candidate[:3] < best[:3]:
            best = candidate

        memo[key] = best
        return best

    return recurse(0, voucher_indices)[3]


def make_upload_names(groups, vouchers):
    upload_names = {}
    assignment_details = []
    seen_missing = defaultdict(int)
    reserved_names = {voucher["path"].stem for voucher in vouchers}

    summaries = [
        {
            "serial": serial,
            "rows": rows,
            "supplier": rows[-1]["supplier"] or rows[0]["supplier"],
            "total": group_total(rows),
        }
        for serial, rows in groups
    ]

    run = []
    for item in summaries + [None]:
        if item is not None and (not run or norm(item["supplier"]) == norm(run[-1]["supplier"])):
            run.append(item)
            continue
        if run:
            plan = solve_supplier_run(run, vouchers)
            for chunk in plan:
                chunk_items = run[chunk["start"] : chunk["end"] + 1]
                last_serial = chunk_items[-1]["serial"]
                upload_name = chunk["upload_name"]
                if chunk["missing"]:
                    upload_name = make_missing_name(chunk["amount"], chunk_items[0]["supplier"], seen_missing, reserved_names)
                else:
                    reserved_names.add(upload_name)
                upload_names[last_serial] = upload_name
                if chunk["voucher_index"] is not None:
                    vouchers[chunk["voucher_index"]]["planned_used"] = True
                assignment_details.append(
                    [
                        ",".join(str(item["serial"]) for item in chunk_items),
                        upload_name,
                        float(chunk["amount"]),
                        chunk_items[0]["supplier"],
                        "缺少支付凭证" if chunk["missing"] else "已分配支付凭证",
                    ]
                )
        run = [] if item is None else [item]

    return upload_names, assignment_details


def build_table(groups, upload_names, assignment_details):
    table_rows = []
    detail_rows = []
    for serial, rows in groups:
        upload_name = upload_names.get(serial, "")
        for index, row in enumerate(rows):
            excel_row = len(table_rows) + 2
            table_rows.append(
                [
                    row["serial"],
                    row["po"],
                    row["supplier"],
                    upload_name if index == len(rows) - 1 else "",
                    float(row["stock_amount"]),
                    row["date"],
                    f"=E{excel_row}+H{excel_row}",
                    row["diff"],
                    row["system_amount"],
                    "待上传",
                    row["remark"],
                ]
            )
        detail_rows.append(
            [
                serial,
                upload_name,
                len(rows),
                float(sum((row["stock_amount"] for row in rows), Decimal("0.00"))),
                rows[0]["supplier"],
                ",".join(str(row["po"]) for row in rows),
            ]
        )
    detail_rows.extend(["", "", "", "", "", ""] for _ in range(1))
    detail_rows.extend(
        [serials, name, "", amount, supplier, status]
        for serials, name, amount, supplier, status in assignment_details
    )
    return table_rows, detail_rows


def parse_vouchers():
    vouchers = []
    for path in sorted(VOUCHER_DIR.glob("*"), key=lambda item: item.name):
        if not path.is_file():
            continue
        match = VOUCHER_RE.match(path.name)
        if not match:
            vouchers.append(
                {
                    "path": path,
                    "amount": None,
                    "supplier": path.stem,
                    "ext": path.suffix.lstrip("."),
                    "used": False,
                    "planned_used": False,
                }
            )
            continue
        vouchers.append(
            {
                "path": path,
                "amount": dec(match.group("amount")),
                "supplier": match.group("supplier"),
                "ext": match.group("ext"),
                "used": False,
                "planned_used": False,
            }
        )
    return vouchers


def supplier_score(source_supplier, target_supplier):
    source = norm(source_supplier)
    target = norm(target_supplier)
    if source == target:
        return 100
    if source in target or target in source:
        return 80
    source_short = source[:2]
    target_short = target[:2]
    if source_short and source_short == target_short:
        return 50
    return 0


def find_voucher(vouchers, upload_name):
    name_match = VOUCHER_RE.match(f"{upload_name}.jpg")
    if not name_match:
        return None
    target_amount = dec(name_match.group("amount"))
    target_supplier = name_match.group("supplier")

    candidates = []
    for voucher in vouchers:
        if voucher["used"] or voucher["amount"] != target_amount:
            continue
        score = supplier_score(voucher["supplier"], target_supplier)
        if score:
            candidates.append((score, len(voucher["supplier"]), voucher))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    return candidates[0][2]


def rename_vouchers(upload_names):
    vouchers = parse_vouchers()
    log = []
    operations = []

    for serial in sorted(upload_names.keys(), key=sort_key):
        upload_name = upload_names[serial]
        voucher = find_voucher(vouchers, upload_name)
        target = VOUCHER_DIR / f"{upload_name}.jpg"
        if voucher is None:
            log.append([serial, "", target.name, "缺少支付凭证"])
            continue
        voucher["used"] = True
        source = voucher["path"]
        if source.name == target.name:
            log.append([serial, source.name, target.name, "无需改名"])
            continue
        temp = source.with_name(f".__tmp_rename_{source.stem}_{os.getpid()}{source.suffix}")
        source.rename(temp)
        operations.append((serial, temp, target, source.name))

    for serial, temp, target, original_name in operations:
        if target.exists():
            log.append([serial, original_name, target.name, "失败：目标文件已存在"])
            continue
        temp.rename(target)
        log.append([serial, original_name, target.name, "已改名"])

    for voucher in vouchers:
        if not voucher["used"] and voucher["path"].exists():
            log.append(["", voucher["path"].name, "", "未被原表序号使用"])

    return log


def style_sheet(worksheet):
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")
    worksheet.freeze_panes = "A2"

    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    for col in range(1, worksheet.max_column + 1):
        width = 10
        for cell in worksheet[get_column_letter(col)]:
            width = max(width, len(str(cell.value)) if cell.value is not None else 0)
        worksheet.column_dimensions[get_column_letter(col)].width = min(width + 2, 48)


def write_workbook(table_rows, detail_rows, rename_log):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["序号", "采购单号", "供应商", "待上传的支付凭证名", "入库金额", "日期", "金额", "差异", "系统采购金额", "状态", "备注"])
    for row in table_rows:
        sheet.append(row)

    detail_sheet = workbook.create_sheet("序号分组明细")
    detail_sheet.append(["序号/序号组合", "待上传的支付凭证名", "采购行数", "分组金额合计", "供应商", "采购单号/状态"])
    for row in detail_rows:
        detail_sheet.append(row)

    rename_sheet = workbook.create_sheet("改名记录")
    rename_sheet.append(["序号", "原文件名", "新文件名", "结果"])
    for row in rename_log:
        rename_sheet.append(row)

    for worksheet in workbook.worksheets:
        style_sheet(worksheet)

    if REFERENCE_XLSX.exists():
        reference = load_workbook(REFERENCE_XLSX)
        reference_sheet = reference.active
        for col in range(1, min(sheet.max_column, reference_sheet.max_column) + 1):
            letter = get_column_letter(col)
            sheet.column_dimensions[letter].width = reference_sheet.column_dimensions[letter].width

    try:
        workbook.save(OUT_XLSX)
    except PermissionError:
        fallback = OUT_XLSX.with_name("6.8报销_影刀格式_修正版.xlsx")
        workbook.save(fallback)
        print(f"OUT_XLSX_LOCKED={OUT_XLSX}")
        print(f"OUT_XLSX_FALLBACK={fallback}")


def write_log(rename_log):
    with LOG_TXT.open("w", encoding="utf-8") as handle:
        for row in rename_log:
            handle.write("\t".join(str(value) for value in row) + "\n")


def main():
    OUT_DIR.mkdir(exist_ok=True)
    groups = load_source_groups()
    planned_vouchers = parse_vouchers()
    upload_names, assignment_details = make_upload_names(groups, planned_vouchers)
    table_rows, detail_rows = build_table(groups, upload_names, assignment_details)
    rename_log = rename_vouchers(upload_names)
    write_workbook(table_rows, detail_rows, rename_log)
    write_log(rename_log)

    print(f"OUT_XLSX={OUT_XLSX}")
    print(f"LOG_TXT={LOG_TXT}")
    print(f"GROUPS={len(groups)}")
    print(f"TABLE_ROWS={len(table_rows)}")
    print(f"RENAMED={sum(1 for row in rename_log if row[3] == '已改名')}")
    print(f"UNCHANGED={sum(1 for row in rename_log if row[3] == '无需改名')}")
    print(f"MISSING={sum(1 for row in rename_log if row[3] == '缺少支付凭证')}")
    print(f"UNUSED={sum(1 for row in rename_log if row[3] == '未被原表序号使用')}")


if __name__ == "__main__":
    main()
