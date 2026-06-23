import openpyxl, shutil, json
from datetime import datetime


def U(s): return s.encode('ascii').decode('unicode_escape')
base = "D:" + "\\" + U("\\u62a5\\u9500\\u5de5\\u4f5c\\u53f0") + "\\04_WorkBuddy" + U("\\u4efb\\u52a1")
path = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868.xlsx")
backup = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868_backup_before_supplier_docs_20260623_") + datetime.now().strftime('%H%M%S') + ".xlsx"

records = [
    # 宝年丰 BNF2606220214，供应商按KG开单，图灵采购单位为米；做表按每条采购单1米，实际单价取供应商行金额，便于影刀录入后金额一致。
    dict(task='WB-IN-20260623-XW-MCG202606220028', order='MCG202606220028', status='待执行', purchase_status='待入库', supplier='宝年丰布业', material='全棉拉架平纹-180G', color='浅紫', color_no='157#', unit='米', qty=1, price=18.7, label='待确认', source='宝年丰布业 BNF2606220214', match='已匹配', note='6032#32S棉拉架 157#浅紫；供应商行金额18.7，图灵按1米录入单价18.7。'),
    dict(task='WB-IN-20260623-XW-MCG202606220020', order='MCG202606220020', status='待执行', purchase_status='待入库', supplier='宝年丰布业', material='全棉拉架平纹-180G', color='姜黄色', color_no='46#', unit='米', qty=1, price=18.7, label='待确认', source='宝年丰布业 BNF2606220214', match='已匹配', note='6032#32S棉拉架 46#金黄/姜黄色；供应商行金额18.7。'),
    dict(task='WB-IN-20260623-XW-MCG202606220019', order='MCG202606220019', status='待执行', purchase_status='待入库', supplier='宝年丰布业', material='全棉拉架平纹-180G', color='红色', color_no='34#', unit='米', qty=1, price=20.2, label='待确认', source='宝年丰布业 BNF2606220214', match='已匹配', note='6032#32S棉拉架 34#大红；供应商行金额20.2。'),
    dict(task='WB-IN-20260623-XW-MCG202606220014', order='MCG202606220014', status='待执行', purchase_status='待入库', supplier='宝年丰布业', material='全棉拉架平纹-180G', color='紫色', color_no='150#', unit='米', qty=1, price=18.7, label='待确认', source='宝年丰布业 BNF2606220214', match='已匹配', note='6032#32S棉拉架 150#紫；供应商行金额18.7。'),
    dict(task='WB-IN-20260623-XW-MCG202606220009', order='MCG202606220009', status='待执行', purchase_status='待入库', supplier='宝年丰布业', material='全棉拉架平纹-180G', color='浅紫', color_no='73#', unit='米', qty=1, price=18.7, label='待确认', source='宝年丰布业 BNF2606220214', match='已匹配', note='6032#32S棉拉架 73#浅紫；供应商行金额18.7。'),
    dict(task='WB-IN-20260623-XW-MCG202606220012', order='MCG202606220012', status='待执行', purchase_status='待入库', supplier='宝年丰布业', material='全棉拉架平纹-180G', color='粉色', color_no='91#', unit='米', qty=1, price=18.7, label='待确认', source='宝年丰布业 BNF2606220214', match='已匹配', note='6032#32S棉拉架 91#粉红；供应商行金额18.7。'),
    dict(task='WB-IN-20260623-XW-MCG202606220027', order='MCG202606220027', status='待执行', purchase_status='待入库', supplier='宝年丰布业', material='全棉拉架平纹-180G', color='白色', color_no='29#', unit='米', qty=1, price=17.2, label='待确认', source='宝年丰布业 BNF2606220214', match='已匹配', note='6032#32S棉拉架 29#蓝光白；供应商行金额17.2。'),
    # 新协纺织单据，本地采购导出供应商名为广新布行，物料/色号/数量唯一对应。
    dict(task='WB-IN-20260623-XW-MCG202606180004', order='MCG202606180004', status='待执行', purchase_status='待入库', supplier='广新布行（新协纺织单据）', material='75D双面佳积布-120G', color='红色', color_no='124#', unit='米', qty=1, price=10, label='待确认', source='新协纺织 XSFH2606180047', match='已匹配', note='供应商单据：75D双面佳织布/佳积布 124#，1米，10元；本地采购单供应商为广新布行。'),
    # 隆发
    dict(task='WB-IN-20260623-XW-MCG202606170037', order='MCG202606170037', status='待执行', purchase_status='待入库', supplier='隆发罗纹', material='全棉2*2拉架罗纹-460G', color='深蓝色', color_no='368', unit='米', qty=1, price=25, label='待确认', source='隆发罗纹 XS202606180009', match='已匹配', note='C版-棉2X2拉架罗纹/2*2棉21支，368深蓝，1米，25元。'),
    # 待匹配，不给影刀执行
    dict(task='WB-IN-20260623-XW-PENDING-BINTAO-260427004', order='', status='待确认', purchase_status='待确认', supplier='佛山市缤涛纺织有限公司', material='BT0015A 阻燃仿棉1*1拉架罗纹', color='大红', color_no='2#大红', unit='米', qty=2, price=13, label='待确认', source='缤涛成品销售出货单 260427004', match='待确认', note='供应商单据：BT0015A 2#大红，2米，13元/米，金额26；本地6月22/18导出未唯一匹配到对应MCG，需补采购单号后给影刀执行。'),
    dict(task='WB-IN-20260623-XW-PENDING-HUAXUN-PL2606220536', order='', status='待确认', purchase_status='待确认', supplier='华讯昌达/易联达联盟', material='仿棉2*2罗纹-RS15#', color='丁香紫', color_no='RS15#', unit='KG', qty=0.9, price=37, label='待确认', source='华讯昌达 STE26062201969\\PL2606220536', match='待确认', note='供应商单据：仿棉2*2罗纹-RS15#丁香紫，0.9KG，37元/KG，金额33；本地导出未唯一匹配到MCG，需补采购单号及图灵单位口径后执行。'),
]

shutil.copy2(path, backup)
wb = openpyxl.load_workbook(path)
ws = wb.worksheets[1]
existing_by_order = {}
existing_by_task = {}
for r in range(2, ws.max_row + 1):
    task = ws.cell(r, 1).value
    order = ws.cell(r, 2).value
    if task:
        existing_by_task[task] = r
    if order:
        existing_by_order[order] = r

added, updated = [], []
for rec in records:
    row_values = [
        rec['task'], rec['order'], rec['purchase_status'], rec['supplier'], rec['material'], rec['color'], rec['color_no'], rec['unit'], '',
        rec['qty'], rec['price'], round(rec['qty'] * rec['price'], 4) if isinstance(rec['qty'], (int, float)) and isinstance(rec['price'], (int, float)) else '',
        rec['label'], '来源供应商单据: ' + rec['source'] + '；' + rec['note'], '供应商单据/本地采购单导出匹配', rec['match'],
        '影刀执行前核对页面明细；待确认记录不得执行。' if rec['status']=='待执行' else '先补采购单号/单位口径，确认后再交影刀执行。', rec['status']
    ]
    target = None
    if rec['order'] and rec['order'] in existing_by_order:
        target = existing_by_order[rec['order']]
    elif rec['task'] in existing_by_task:
        target = existing_by_task[rec['task']]
    if target:
        for c, v in enumerate(row_values, start=1):
            ws.cell(target, c).value = v
        updated.append(rec['order'] or rec['task'])
    else:
        ws.append(row_values)
        added.append(rec['order'] or rec['task'])

wb.save(path)
print(json.dumps({'xlsx': path, 'backup': backup, 'added': added, 'updated': updated, 'total_records': len(records)}, ensure_ascii=False, indent=2))
