import openpyxl, shutil, json
from datetime import datetime

def U(s): return s.encode('ascii').decode('unicode_escape')
base = "D:" + "\\" + U("\\u62a5\\u9500\\u5de5\\u4f5c\\u53f0") + "\\04_WorkBuddy" + U("\\u4efb\\u52a1")
path = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868.xlsx")
backup = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868_backup_before_yifeng_price_20260622_") + datetime.now().strftime('%H%M%S') + ".xlsx"
prices = {
    'MCG202606180003': 30,
    'MCG202606160044': 30,
    'MCG202606180024': 20,
    'MCG202606180015': 20,
    'MCG202606180001': 20,
    'MCG202606170001': 20,
}
raw = {
    'MCG202606180003': '棉涤卫衣（无拉架）-300G 米色，1米，用户确认怡丰实价30元/米',
    'MCG202606160044': '棉涤卫衣（无拉架）-300G 黑色，1米，用户确认怡丰实价30元/米',
    'MCG202606180024': '纯棉单面（无拉架）花灰2，1米，用户确认怡丰实价20元/米',
    'MCG202606180015': '全棉拉架平纹-190G 深蓝色，1米，用户确认怡丰实价20元/米',
    'MCG202606180001': '棉1*1拉架罗纹-180G 浅蓝，1米，用户确认怡丰实价20元/米',
    'MCG202606170001': '棉1*1拉架罗纹-180G 粉色，1米，用户确认怡丰实价20元/米',
}
shutil.copy2(path, backup)
wb = openpyxl.load_workbook(path)
ws = wb.worksheets[1]
updated=[]
for r in range(2, ws.max_row + 1):
    order = ws.cell(r, 2).value
    if order in prices:
        price = prices[order]
        ws.cell(r, 3).value = '待入库'
        ws.cell(r, 4).value = '怡丰布行'
        ws.cell(r, 8).value = '米'
        ws.cell(r, 10).value = 1
        ws.cell(r, 11).value = price
        ws.cell(r, 12).value = price
        ws.cell(r, 13).value = '待确认'
        ws.cell(r, 14).value = '来源供应商单据: 怡丰布行；' + raw[order] + '；差异原因：部分物料单价优惠，用户确认可按此价格执行。'
        ws.cell(r, 15).value = '供应商单据/用户确认'
        ws.cell(r, 16).value = '已匹配'
        ws.cell(r, 17).value = '小王已补齐怡丰数量单价，交小吴入库前仍需核对页面明细。'
        ws.cell(r, 18).value = '待执行'
        updated.append(order)
wb.save(path)
print(json.dumps({'xlsx': path, 'backup': backup, 'updated': updated}, ensure_ascii=False, indent=2))
