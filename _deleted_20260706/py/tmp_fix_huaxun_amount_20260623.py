import openpyxl, shutil, json
from datetime import datetime

def U(s): return s.encode('ascii').decode('unicode_escape')
base = "D:" + "\\" + U("\\u62a5\\u9500\\u5de5\\u4f5c\\u53f0") + "\\04_WorkBuddy" + U("\\u4efb\\u52a1")
path = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868.xlsx")
backup = base + "\\WorkBuddy" + U("\\u91c7\\u8d2d\\u5165\\u5e93\\u4efb\\u52a1\\u8868_backup_before_huaxun_amount_20260623_") + datetime.now().strftime('%H%M%S') + ".xlsx"
shutil.copy2(path, backup)
wb=openpyxl.load_workbook(path)
ws=wb.worksheets[1]
updated=[]
for r in range(2, ws.max_row+1):
    if ws.cell(r,1).value == 'WB-IN-20260623-XW-PENDING-HUAXUN-PL2606220536':
        ws.cell(r,12).value = 33
        ws.cell(r,14).value = '来源供应商单据: 华讯昌达 STE26062201969\\PL2606220536；供应商单据：仿棉2*2罗纹-RS15#丁香紫，0.9KG，37元/KG，单据金额33元（供应商取整）；本地导出未唯一匹配到MCG，需补采购单号及图灵单位口径后执行。'
        updated.append(r)
wb.save(path)
print(json.dumps({'backup': backup, 'updated_rows': updated}, ensure_ascii=False, indent=2))
